# a) A listagem dos vendedores seguido da soma de suas comissões
# b) Apresente o vendedor que mais ganhou comissões
# c) Apresente o vendedor que mais ganhou comissões no mês de Junho

# from openpyxl import load_workbook
#
#
# arquivo = load_workbook("Exercício de Algoritmos Controle de Vendas.xlsx", data_only=True)
# planilha = arquivo.active
# vendedores = []
# for i in range(5,23):
#     conteudo = planilha[f'F{i}']
#     vendedores.append(conteudo.value)
# nv = len(vendedores)
# for j in range(nv):
#     while vendedores[j] == vendedores[-j]:
#         del vendedores[j]
#
#
# print(vendedores)

import openpyxl

workbook = openpyxl.load_workbook(r"Exercício de Algoritmos Controle de Vendas.xlsx")
sheetPlanilha = workbook['Página1']

vendas = {}
vendasJunho = {}

for linha in sheetPlanilha.iter_rows(min_row=5, max_row=22, min_col=1, max_col=8):
    vendedor = sheetPlanilha.cell(row=linha[0].row, column=6).value
    comissao = sheetPlanilha.cell(row=linha[0].row, column=5).value * sheetPlanilha.cell(row=linha[0].row,
                                                                                         column=7).value
    dataVenda = sheetPlanilha.cell(row=linha[0].row, column=1).value

    if vendedor not in vendas:
        vendas[vendedor] = {'Comissão': 0}

    vendas[vendedor]["Comissão"] += comissao

    if dataVenda.month == 6:
        if vendedor not in vendasJunho:
            vendasJunho[vendedor] = {"Comissão": 0}

        vendasJunho[vendedor]["Comissão"] += comissao

for vendedor, valor in vendas.items():
    print(f"Vendedor: {vendedor}, Comissão: R$ {valor['Comissão']:.2f}")

vendedorMaiorComissao = None
maiorComissao = 0

for vendedor, valor in vendas.items():
    if valor['Comissão'] > maiorComissao:
        maiorComissao = valor['Comissão']
        vendedorMaiorComissao = vendedor

print(f'O vendedor com a maior comissão é {vendedorMaiorComissao} com uma comissão total de R$ {maiorComissao:.2f}')

vendedorMaiorComissaojunho = None
maiorComissaojunho = 0

for vendedor, valor in vendasJunho.items():
    if valor['Comissão'] > maiorComissaojunho:
        maiorComissaojunho = valor['Comissão']
        vendedorMaiorComissaojunho = vendedor

print(
    f'O vendedor com maior comissão no mês de Junho é {vendedorMaiorComissaojunho} com comissão totral de R$ {maiorComissaojunho:.2f}')