from openpyxl import load_workbook

#https://letscode.com.br/blog/aprenda-a-integrar-python-e-excel
#https://openpyxl.readthedocs.io/en/stable/tutorial.html

#caminho = 'AULAS-EAD-REGISTRO-DE-ATIVIDADE.xlsx'
caminho = 'Aprendendoalerplanilhas.xlsx'
arquivo_excel = load_workbook(caminho, data_only=True)

planilha = arquivo_excel.active
#conteudo = planilha['B5']
#celula = []
#celula.append(conteudo.value)

# planilha.cell(row=7, column=4, value='')
# arquivo_excel.save("Aprendendoalerplanilhas.xlsx")
alunos = []
for i in range(6, 16):
    conteudo = planilha[f'B{i}']
    alunos.append(conteudo.value)

resultadoAlunosSim = {}
resultadoAlunosNao = {}

for a in alunos:
    somasim = 0
    somanao = 0
    for i in range(6, 45):
        celulaAluno = planilha[f'B{i}']
        celulaResposta = planilha[f'C{i}']
        if celulaResposta.value != None:
            if celulaAluno.value == a and celulaResposta.value == 'SIM':
                somasim += 1
                resultadoAlunosSim[a] = somasim
            elif celulaAluno.value == a and (celulaResposta.value == 'NÃO' or celulaResposta.value == 'NAO'):
                somanao += 1
                resultadoAlunosNao[a] = somanao
print("SIM", resultadoAlunosSim)
print("NÃO", resultadoAlunosNao)


indice = 0
for z in range(49, 49 + len(alunos)):
        planilha.cell(row=z, column=5, value=alunos[indice])
        indice += 1

arquivo_excel.save("Aprendendoalerplanilhas.xlsx")

indice = 0
for z in range(49, 49 + len(alunos)):
    if alunos[indice] == planilha[f'E{z}'].value:
        planilha.cell(row=z, column=6, value=resultadoAlunosSim.get(alunos[indice]))
        planilha.cell(row=z, column=7, value=resultadoAlunosNao.get(alunos[indice]))
        indice += 1

arquivo_excel.save("Aprendendoalerplanilhas.xlsx")